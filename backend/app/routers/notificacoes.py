from fastapi import APIRouter, Depends, HTTPException, status, Response, BackgroundTasks
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, and_, or_
from typing import List, Optional
from datetime import date, datetime, timedelta
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..database import get_db
from ..models import (
    TemplateNotificacao, NotificacaoEnviada, ConfiguracaoNotificacao,
    TipoNotificacao, StatusNotificacao, CanalNotificacao,
    Usuario, Evento, Empresa
)
from ..schemas import (
    TemplateNotificacaoCreate, TemplateNotificacaoUpdate, TemplateNotificacao as TemplateSchema,
    NotificacaoEnviadaResponse, EnviarNotificacaoManual, FiltrosNotificacoes,
    ConfiguracaoNotificacaoUpdate, DashboardNotificacoes
)
from ..auth import obter_usuario_atual, verificar_permissao_admin
from ..services.notification_service import notification_service

router = APIRouter(prefix="/notificacoes", tags=["Notificações Inteligentes"])

@router.get("/templates", response_model=List[TemplateSchema])
async def listar_templates(
    tipo_notificacao: Optional[str] = None,
    canal: Optional[str] = None,
    ativo: Optional[bool] = None,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_admin)
):
    """Listar templates de notificação (apenas admin)"""
    
    query = db.query(TemplateNotificacao).filter(
        TemplateNotificacao.empresa_id == usuario_atual.empresa_id
    )
    
    if tipo_notificacao:
        query = query.filter(TemplateNotificacao.tipo_notificacao == tipo_notificacao)
    if canal:
        query = query.filter(TemplateNotificacao.canal == canal)
    if ativo is not None:
        query = query.filter(TemplateNotificacao.ativo == ativo)
    
    templates = query.order_by(desc(TemplateNotificacao.criado_em)).all()
    return templates

@router.post("/templates", response_model=TemplateSchema)
async def criar_template(
    template: TemplateNotificacaoCreate,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_admin)
):
    """Criar novo template de notificação (apenas admin)"""
    
    existente = db.query(TemplateNotificacao).filter(
        TemplateNotificacao.empresa_id == usuario_atual.empresa_id,
        TemplateNotificacao.tipo_notificacao == template.tipo_notificacao,
        TemplateNotificacao.canal == template.canal,
        TemplateNotificacao.ativo == True
    ).first()
    
    if existente:
        raise HTTPException(
            status_code=400,
            detail="Já existe um template ativo para este tipo/canal"
        )
    
    variaveis = _obter_variaveis_template(template.tipo_notificacao)
    
    db_template = TemplateNotificacao(
        **template.dict(),
        empresa_id=usuario_atual.empresa_id,
        criado_por_id=usuario_atual.id,
        variaveis_disponiveis=variaveis
    )
    
    db.add(db_template)
    db.commit()
    db.refresh(db_template)
    
    return db_template

@router.put("/templates/{template_id}", response_model=TemplateSchema)
async def atualizar_template(
    template_id: int,
    template_update: TemplateNotificacaoUpdate,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_admin)
):
    """Atualizar template de notificação (apenas admin)"""
    
    template = db.query(TemplateNotificacao).filter(
        TemplateNotificacao.id == template_id,
        TemplateNotificacao.empresa_id == usuario_atual.empresa_id
    ).first()
    
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    for field, value in template_update.dict(exclude_unset=True).items():
        setattr(template, field, value)
    
    db.commit()
    db.refresh(template)
    
    return template

@router.delete("/templates/{template_id}")
async def deletar_template(
    template_id: int,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_admin)
):
    """Deletar template de notificação (apenas admin)"""
    
    template = db.query(TemplateNotificacao).filter(
        TemplateNotificacao.id == template_id,
        TemplateNotificacao.empresa_id == usuario_atual.empresa_id
    ).first()
    
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    template.ativo = False
    db.commit()
    
    return {"message": "Template desativado com sucesso"}

@router.get("/historico", response_model=List[NotificacaoEnviadaResponse])
async def obter_historico_notificacoes(
    filtros: FiltrosNotificacoes = Depends(),
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Obter histórico de notificações enviadas"""
    
    query = db.query(
        NotificacaoEnviada.id,
        NotificacaoEnviada.tipo_notificacao,
        NotificacaoEnviada.canal,
        NotificacaoEnviada.destinatario,
        NotificacaoEnviada.titulo,
        NotificacaoEnviada.conteudo,
        NotificacaoEnviada.status,
        NotificacaoEnviada.tentativas,
        NotificacaoEnviada.enviada_em,
        NotificacaoEnviada.criada_em,
        NotificacaoEnviada.erro_detalhes,
        Evento.nome.label('evento_nome'),
        Usuario.nome.label('usuario_nome')
    ).outerjoin(
        Evento, NotificacaoEnviada.evento_id == Evento.id
    ).outerjoin(
        Usuario, NotificacaoEnviada.usuario_id == Usuario.id
    )
    
    if usuario_atual.tipo.value != "admin":
        query = query.join(TemplateNotificacao).filter(
            TemplateNotificacao.empresa_id == usuario_atual.empresa_id
        )
    
    if filtros.tipo_notificacao:
        query = query.filter(NotificacaoEnviada.tipo_notificacao == filtros.tipo_notificacao)
    if filtros.canal:
        query = query.filter(NotificacaoEnviada.canal == filtros.canal)
    if filtros.status:
        query = query.filter(NotificacaoEnviada.status == filtros.status)
    if filtros.evento_id:
        query = query.filter(NotificacaoEnviada.evento_id == filtros.evento_id)
    if filtros.destinatario:
        query = query.filter(NotificacaoEnviada.destinatario.ilike(f"%{filtros.destinatario}%"))
    if filtros.data_inicio:
        query = query.filter(func.date(NotificacaoEnviada.criada_em) >= filtros.data_inicio)
    if filtros.data_fim:
        query = query.filter(func.date(NotificacaoEnviada.criada_em) <= filtros.data_fim)
    
    notificacoes = query.order_by(
        desc(NotificacaoEnviada.criada_em)
    ).offset(filtros.offset).limit(filtros.limit).all()
    
    return [
        NotificacaoEnviadaResponse(
            id=n.id,
            tipo_notificacao=n.tipo_notificacao.value,
            canal=n.canal.value,
            destinatario=n.destinatario,
            titulo=n.titulo,
            conteudo=n.conteudo,
            status=n.status.value,
            tentativas=n.tentativas,
            evento_nome=n.evento_nome,
            usuario_nome=n.usuario_nome,
            enviada_em=n.enviada_em,
            criada_em=n.criada_em,
            erro_detalhes=n.erro_detalhes
        ) for n in notificacoes
    ]

@router.get("/dashboard", response_model=DashboardNotificacoes)
async def obter_dashboard_notificacoes(
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Dashboard de notificações"""
    
    hoje = date.today()
    
    base_query = db.query(NotificacaoEnviada)
    if usuario_atual.tipo.value != "admin":
        base_query = base_query.join(TemplateNotificacao).filter(
            TemplateNotificacao.empresa_id == usuario_atual.empresa_id
        )
    
    total_hoje = base_query.filter(
        func.date(NotificacaoEnviada.criada_em) == hoje
    ).count()
    
    total_pendentes = base_query.filter(
        NotificacaoEnviada.status == StatusNotificacao.PENDENTE
    ).count()
    
    total_falhadas = base_query.filter(
        NotificacaoEnviada.status == StatusNotificacao.FALHADA,
        func.date(NotificacaoEnviada.criada_em) == hoje
    ).count()
    
    total_enviadas = base_query.filter(
        NotificacaoEnviada.status == StatusNotificacao.ENVIADA,
        func.date(NotificacaoEnviada.criada_em) == hoje
    ).count()
    
    taxa_sucesso = (total_enviadas / total_hoje * 100) if total_hoje > 0 else 0
    
    notificacoes_recentes = base_query.order_by(
        desc(NotificacaoEnviada.criada_em)
    ).limit(10).all()
    
    tipos_stats = db.query(
        NotificacaoEnviada.tipo_notificacao,
        func.count(NotificacaoEnviada.id).label('total')
    ).filter(
        func.date(NotificacaoEnviada.criada_em) >= hoje - timedelta(days=7)
    ).group_by(NotificacaoEnviada.tipo_notificacao).all()
    
    canais_stats = db.query(
        NotificacaoEnviada.canal,
        func.count(NotificacaoEnviada.id).label('total'),
        func.count(NotificacaoEnviada.id).filter(NotificacaoEnviada.status == StatusNotificacao.ENVIADA).label('enviadas')
    ).filter(
        func.date(NotificacaoEnviada.criada_em) >= hoje - timedelta(days=7)
    ).group_by(NotificacaoEnviada.canal).all()
    
    return DashboardNotificacoes(
        total_enviadas_hoje=total_enviadas,
        total_pendentes=total_pendentes,
        total_falhadas=total_falhadas,
        taxa_sucesso=round(taxa_sucesso, 2),
        notificacoes_recentes=[
            NotificacaoEnviadaResponse(
                id=n.id,
                tipo_notificacao=n.tipo_notificacao.value,
                canal=n.canal.value,
                destinatario=n.destinatario,
                titulo=n.titulo,
                conteudo=n.conteudo[:100] + "..." if len(n.conteudo) > 100 else n.conteudo,
                status=n.status.value,
                tentativas=n.tentativas,
                enviada_em=n.enviada_em,
                criada_em=n.criada_em
            ) for n in notificacoes_recentes
        ],
        tipos_mais_enviados=[
            {"tipo": t.tipo_notificacao.value, "total": t.total} for t in tipos_stats
        ],
        canais_estatisticas=[
            {
                "canal": c.canal.value,
                "total": c.total,
                "enviadas": c.enviadas or 0,
                "taxa_sucesso": round((c.enviadas or 0) / c.total * 100, 2) if c.total > 0 else 0
            } for c in canais_stats
        ]
    )

@router.post("/enviar-manual")
async def enviar_notificacao_manual(
    dados: EnviarNotificacaoManual,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_admin)
):
    """Enviar notificação manual (apenas admin)"""
    
    background_tasks.add_task(
        notification_service.enviar_notificacao_manual,
        dados.dict(),
        db
    )
    
    return {"message": "Notificação sendo enviada"}

@router.get("/export/{formato}")
async def exportar_notificacoes(
    formato: str,
    filtros: FiltrosNotificacoes = Depends(),
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Exportar histórico de notificações"""
    
    if formato not in ["excel", "csv"]:
        raise HTTPException(status_code=400, detail="Formato não suportado")
    
    filtros.limit = 10000
    notificacoes = await obter_historico_notificacoes(filtros, db, usuario_atual)
    
    if formato == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico Notificações"
        
        headers = [
            'ID', 'Tipo', 'Canal', 'Destinatário', 'Título', 'Status',
            'Tentativas', 'Evento', 'Enviada em', 'Criada em'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for row, notif in enumerate(notificacoes, 2):
            ws.cell(row=row, column=1, value=notif.id)
            ws.cell(row=row, column=2, value=notif.tipo_notificacao)
            ws.cell(row=row, column=3, value=notif.canal)
            ws.cell(row=row, column=4, value=notif.destinatario)
            ws.cell(row=row, column=5, value=notif.titulo)
            ws.cell(row=row, column=6, value=notif.status)
            ws.cell(row=row, column=7, value=notif.tentativas)
            ws.cell(row=row, column=8, value=notif.evento_nome)
            ws.cell(row=row, column=9, value=notif.enviada_em.strftime("%d/%m/%Y %H:%M") if notif.enviada_em else "")
            ws.cell(row=row, column=10, value=notif.criada_em.strftime("%d/%m/%Y %H:%M"))
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=notificacoes.xlsx"}
        )
    
    elif formato == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow([
            'ID', 'Tipo', 'Canal', 'Destinatário', 'Título', 'Status',
            'Tentativas', 'Evento', 'Enviada em', 'Criada em'
        ])
        
        for notif in notificacoes:
            writer.writerow([
                notif.id,
                notif.tipo_notificacao,
                notif.canal,
                notif.destinatario,
                notif.titulo,
                notif.status,
                notif.tentativas,
                notif.evento_nome,
                notif.enviada_em.strftime("%d/%m/%Y %H:%M") if notif.enviada_em else "",
                notif.criada_em.strftime("%d/%m/%Y %H:%M")
            ])
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=notificacoes.csv"}
        )

@router.get("/configuracoes")
async def obter_configuracoes(
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_admin)
):
    """Obter configurações de notificação (apenas admin)"""
    
    config = db.query(ConfiguracaoNotificacao).filter(
        ConfiguracaoNotificacao.empresa_id == usuario_atual.empresa_id
    ).first()
    
    if not config:
        config = ConfiguracaoNotificacao(
            empresa_id=usuario_atual.empresa_id,
            whatsapp_ativo=True,
            sms_ativo=False,
            email_ativo=False
        )
        db.add(config)
        db.commit()
        db.refresh(config)
    
    return config

@router.put("/configuracoes")
async def atualizar_configuracoes(
    config_update: ConfiguracaoNotificacaoUpdate,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_admin)
):
    """Atualizar configurações de notificação (apenas admin)"""
    
    config = db.query(ConfiguracaoNotificacao).filter(
        ConfiguracaoNotificacao.empresa_id == usuario_atual.empresa_id
    ).first()
    
    if not config:
        config = ConfiguracaoNotificacao(empresa_id=usuario_atual.empresa_id)
        db.add(config)
    
    for field, value in config_update.dict(exclude_unset=True).items():
        setattr(config, field, value)
    
    db.commit()
    db.refresh(config)
    
    return {"message": "Configurações atualizadas com sucesso"}

@router.get("/tipos-disponiveis")
async def obter_tipos_notificacao():
    """Obter tipos de notificação disponíveis"""
    return [
        {
            "value": tipo.value,
            "label": _obter_label_tipo(tipo),
            "variaveis": _obter_variaveis_template(tipo.value)
        } for tipo in TipoNotificacao
    ]

@router.get("/canais-disponiveis")
async def obter_canais_notificacao():
    """Obter canais de notificação disponíveis"""
    return [
        {"value": canal.value, "label": _obter_label_canal(canal)}
        for canal in CanalNotificacao
    ]

@router.post("/testar-canal/{canal}")
async def testar_canal_notificacao(
    canal: str,
    destinatario: str,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_admin)
):
    """Testar envio de notificação por canal específico (apenas admin)"""
    
    if canal not in [c.value for c in CanalNotificacao]:
        raise HTTPException(status_code=400, detail="Canal inválido")
    
    notificacao_teste = NotificacaoEnviada(
        tipo_notificacao=TipoNotificacao.VENDA_CONFIRMADA,
        canal=CanalNotificacao(canal),
        destinatario=destinatario,
        titulo="🧪 Teste de Notificação",
        conteudo=f"Esta é uma mensagem de teste do canal {canal.upper()}. Sistema funcionando corretamente!",
        dados_contexto='{"teste": true}'
    )
    
    db.add(notificacao_teste)
    db.commit()
    db.refresh(notificacao_teste)
    
    try:
        await notification_service._enviar_notificacao(notificacao_teste, db)
        return {
            "sucesso": True,
            "mensagem": f"Teste de {canal} enviado com sucesso",
            "notificacao_id": notificacao_teste.id,
            "status": notificacao_teste.status.value
        }
    except Exception as e:
        return {
            "sucesso": False,
            "mensagem": f"Erro no teste de {canal}: {str(e)}",
            "notificacao_id": notificacao_teste.id,
            "erro": str(e)
        }

@router.get("/estatisticas-canais")
async def obter_estatisticas_canais(
    periodo_dias: int = 30,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Obter estatísticas detalhadas por canal"""
    
    data_inicio = datetime.now() - timedelta(days=periodo_dias)
    
    base_query = db.query(NotificacaoEnviada).filter(
        NotificacaoEnviada.criada_em >= data_inicio
    )
    
    if usuario_atual.tipo.value != "admin":
        base_query = base_query.join(TemplateNotificacao).filter(
            TemplateNotificacao.empresa_id == usuario_atual.empresa_id
        )
    
    stats_por_canal = {}
    
    for canal in CanalNotificacao:
        canal_query = base_query.filter(NotificacaoEnviada.canal == canal)
        
        total = canal_query.count()
        enviadas = canal_query.filter(NotificacaoEnviada.status == StatusNotificacao.ENVIADA).count()
        falhadas = canal_query.filter(NotificacaoEnviada.status == StatusNotificacao.FALHADA).count()
        pendentes = canal_query.filter(NotificacaoEnviada.status == StatusNotificacao.PENDENTE).count()
        
        tempo_medio = db.query(
            func.avg(
                func.extract('epoch', NotificacaoEnviada.enviada_em) - 
                func.extract('epoch', NotificacaoEnviada.criada_em)
            )
        ).filter(
            NotificacaoEnviada.canal == canal,
            NotificacaoEnviada.status == StatusNotificacao.ENVIADA,
            NotificacaoEnviada.criada_em >= data_inicio
        ).scalar() or 0
        
        stats_por_canal[canal.value] = {
            "canal": _obter_label_canal(canal),
            "total": total,
            "enviadas": enviadas,
            "falhadas": falhadas,
            "pendentes": pendentes,
            "taxa_sucesso": round((enviadas / total * 100) if total > 0 else 0, 2),
            "tempo_medio_segundos": round(tempo_medio, 2),
            "ativo": total > 0
        }
    
    return {
        "periodo_dias": periodo_dias,
        "data_inicio": data_inicio.isoformat(),
        "estatisticas": stats_por_canal,
        "resumo": {
            "canal_mais_usado": max(stats_por_canal.items(), key=lambda x: x[1]["total"])[0] if any(s["total"] > 0 for s in stats_por_canal.values()) else None,
            "melhor_taxa_sucesso": max(stats_por_canal.items(), key=lambda x: x[1]["taxa_sucesso"])[0] if any(s["total"] > 0 for s in stats_por_canal.values()) else None,
            "total_geral": sum(s["total"] for s in stats_por_canal.values())
        }
    }

def _obter_variaveis_template(tipo_notificacao: str) -> str:
    """Obter variáveis disponíveis para template"""
    variaveis_base = ["nome", "data_atual", "hora_atual"]
    
    variaveis_especificas = {
        "venda_confirmada": ["evento_nome", "data_evento", "local_evento", "valor", "lista_nome"],
        "checkin_realizado": ["evento_nome", "data_evento", "local_evento"],
        "caixa_fechado": ["evento_nome", "receita_total", "total_vendas"],
        "aniversariante": ["evento_nome", "data_evento"],
        "conquista_desbloqueada": ["conquista_nome", "badge_nivel"],
        "ranking_atualizado": ["posicao_ranking", "total_vendas", "receita_total"],
        "evento_criado": ["evento_nome", "data_evento", "local_evento"],
        "lista_criada": ["lista_nome", "evento_nome"]
    }
    
    variaveis = variaveis_base + variaveis_especificas.get(tipo_notificacao, [])
    return ", ".join([f"{{{var}}}" for var in variaveis])

def _obter_label_tipo(tipo: TipoNotificacao) -> str:
    """Obter label amigável para tipo de notificação"""
    labels = {
        TipoNotificacao.VENDA_CONFIRMADA: "Venda Confirmada",
        TipoNotificacao.CHECKIN_REALIZADO: "Check-in Realizado",
        TipoNotificacao.CAIXA_FECHADO: "Caixa Fechado",
        TipoNotificacao.ALERTA_FINANCEIRO: "Alerta Financeiro",
        TipoNotificacao.ANIVERSARIANTE: "Aniversariante",
        TipoNotificacao.RANKING_ATUALIZADO: "Ranking Atualizado",
        TipoNotificacao.CONQUISTA_DESBLOQUEADA: "Conquista Desbloqueada",
        TipoNotificacao.EVENTO_CRIADO: "Evento Criado",
        TipoNotificacao.LISTA_CRIADA: "Lista Criada"
    }
    return labels.get(tipo, tipo.value)

def _obter_label_canal(canal: CanalNotificacao) -> str:
    """Obter label amigável para canal de notificação"""
    labels = {
        CanalNotificacao.WHATSAPP: "WhatsApp",
        CanalNotificacao.SMS: "SMS",
        CanalNotificacao.EMAIL: "E-mail",
        CanalNotificacao.PUSH: "Push Notification"
    }
    return labels.get(canal, canal.value)
